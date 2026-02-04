from django.shortcuts import render,redirect, get_object_or_404
from django.contrib import messages  
from django.contrib.auth.decorators import login_required ,permission_required
from django.contrib.auth import login, logout 
from .form import RegisterForm, AppSettingsForm, EditUserForm, DepartmentForm, GroupForm
from .models import AppSettings ,User, UserLoginSession, Department
import openpyxl
from django.http import HttpResponse 
from django.contrib.auth import get_user_model, update_session_auth_hash
from django.core.exceptions import ValidationError
from django.contrib.auth.password_validation import validate_password
from django.contrib.auth.models import Group, Permission
from django.db.models import OuterRef, Subquery
from django.contrib.sessions.models import Session  
from django.http import JsonResponse

@login_required
def close_session(request, session_id):
    session = get_object_or_404(
        UserLoginSession,
        id=session_id,
        user=request.user  # ปลอดภัย: ปิดได้เฉพาะของตัวเอง
    ) 
    # ลบ session จริง
    Session.objects.filter(session_key=session.session_key).delete()

    # ลบ record
    session.delete()

    return redirect('profile')  # หรือชื่อหน้า profile ของคุณ

def export_users_excel(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Users'

    # Header
    sheet.append(['ID', 'Username', 'Email', 'first Name', 'Phone'])

    for user in User.objects.all():
        sheet.append([user.id, user.username, user.email, user.first_name, user.phone])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=users.xlsx'
    workbook.save(response) 
    return response
 
User = get_user_model() 
def import_users_excel(request):
    if request.method == 'POST' and request.FILES.get('file'):
        excel_file = request.FILES['file']

        # โหลด workbook
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        # อ่าน header ของ Excel
        headers = [cell.value for cell in sheet[1]]  # แถวแรก

        for row in sheet.iter_rows(min_row=2, values_only=True):  # เริ่มจากแถว 2
            data = dict(zip(headers, row))

            # Update หรือ Create ตาม username
            if 'Username' in data:
                User.objects.update_or_create(
                    username=data['Username'],
                    defaults={
                        'email': data.get('Email', ''),
                        'first_name': data.get('first Name', ''),
                        'phone': data.get('Phone', ''),
                    }
                ) 
        messages.success(request, "Users imported successfully from Excel!")
        return redirect('user_list') 
    messages.error(request, "No file uploaded.")
    return redirect('user_list')
# Create your views here.
@login_required
def profile(request):
    users = User.objects.all()
    sessions = UserLoginSession.objects.filter(user=request.user)
    return render(request, 'profile.html', {'users': users,"user_sessions": sessions}) 

def logout_view(request):
    logout(request)
    return redirect('login')

def settings_view(request):
    settings = AppSettings.get_settings()
    if request.method == "POST":
        form = AppSettingsForm(request.POST, request.FILES, instance=settings)
        if form.is_valid():  
            # Remove favicon if flagged
            if request.POST.get('remove_favicon') == '1' and settings.favicon:
                settings.favicon.delete(save=False)
                settings.favicon = None
            if request.POST.get('remove_app_icon') == '1' and settings.app_icon:
                settings.app_icon.delete(save=False)
                settings.app_icon = None 
            form.save()
        return redirect("settings")  
    else:
        form = AppSettingsForm(instance=settings) 
    return render(request, "setting.html", {"form": form})

@login_required
def dashboard(request): 
    total_users = User.objects.count()  
    # Orders per month (last 12 months)  
    context = { 
        'total_users': total_users,  
    }
    return render(request, 'dashboard.html', context) 

@login_required
def user_list(request): 
    users = User.objects.all().order_by('id')
    create_form = RegisterForm()  # empty form for Add modal 
    last_sessions = UserLoginSession.objects.filter(
        user=OuterRef('pk')
    ).order_by('-last_activity') 
    users = users.annotate(
        last_device=Subquery(last_sessions.values('device')[:1]),
        last_browser=Subquery(last_sessions.values('browser')[:1]),
        last_os=Subquery(last_sessions.values('os')[:1]),
        last_ip=Subquery(last_sessions.values('ip_address')[:1]),
        last_activity=Subquery(last_sessions.values('last_activity')[:1]),
    ) 
    user_with_forms = [(user, EditUserForm(instance=user)) for user in users] 
    return render(request, 'user_list.html', {'users': users, 'create_form': create_form, 'user_with_forms': user_with_forms})
  
    
def register_view(request):  
    if request.method == 'POST':  
        form = RegisterForm(request.POST, request.FILES)        
        if form.is_valid():
            user = form.save(commit=False)  
            # 🔹 Handle new avatar upload
            if 'avatar' in request.FILES:
                user.avatar = request.FILES['avatar']
            user.save() 
            # 🔥 สำคัญมาก
            user.save_m2m()
            login(request, user)
            return redirect('dashboard')
    else:
        form = RegisterForm()
    return render(request, 'register.html', {'form': form })
def register_on_system(request):  
    if request.method == 'POST':  
        form = RegisterForm(request.POST, request.FILES)      
        if form.is_valid():
            user = form.save(commit=False)   
            if 'avatar' in request.FILES:
                user.avatar = request.FILES['avatar']
            user.save()  
            return redirect('user_list')
    
@login_required
def edit_user(request, user_id):
    user_obj = get_object_or_404(User, pk=user_id)
    if request.method == 'POST':
        form = EditUserForm(request.POST, request.FILES, instance=user_obj)
        if form.is_valid():
            user = form.save(commit=False)

            # Remove avatar if flagged
            if request.POST.get('remove_avatar') == 'true' and user.avatar:
                user.avatar.delete(save=False)
                user.avatar = None

            # Assign new avatar if uploaded
            if 'avatar' in request.FILES:
                user.avatar = request.FILES['avatar']

            user.save()  # save all fields + avatar changes

            # 🔥 สำคัญมาก
            form.save_m2m()
        return redirect('user_list')
    
@login_required
def delete_user(request, user_id):
    user = get_object_or_404(User, id=user_id) 
    user.delete()
    return redirect('user_list')
    
@login_required
def change_password(request, user_id):
    user_obj = get_object_or_404(User, pk=user_id)
    if request.method == 'POST':
        password = request.POST.get('password') 
        if not password:
            messages.error(request, "Both fields are required.")
            return redirect(request.META.get('HTTP_REFERER', 'user_list'))
        try:
            validate_password(password, user_obj)
        except ValidationError as e:
            messages.error(request, e.messages[0])
            return redirect(request.META.get('HTTP_REFERER', 'user_list'))
        user_obj.set_password(password)
        user_obj.save()
        # Update session if user changes own password
        if request.user == user_obj:
            update_session_auth_hash(request, user_obj)
        messages.success(request, f"Password updated for {user_obj.username}.")
        return redirect(request.META.get('HTTP_REFERER', 'user_list'))
    
def group_list(request):
    groups = Group.objects.all().order_by('id')
    create_form = GroupForm()  # for Add modal 
    # Prepare edit forms per group instance
    # สร้าง tuple (group, form)
    group_with_forms = [(group, GroupForm(instance=group)) for group in groups]
    return render(request, 'group_list.html', {
        'groups': groups,
        'create_form': create_form,
        'group_with_forms': group_with_forms,  # send to template
    })
    
def group_create(request):
    if request.method == 'POST':
        form = GroupForm(request.POST)
        if form.is_valid():
            group = form.save(commit=False)
            group.save()
            group.permissions.set(request.POST.getlist('permissions'))
    return redirect('group_list')

def group_edit(request, group_id):
    group = get_object_or_404(Group, pk=group_id)
    if request.method == 'POST':
        form = GroupForm(request.POST, instance=group)
        if form.is_valid():
            group = form.save(commit=False)
            group.save()
            group.permissions.set(request.POST.getlist('permissions'))
    return redirect('group_list')

def group_delete(request, group_id):
    group = get_object_or_404(Group, pk=group_id)
    group.delete()
    return redirect('group_list')

def department_list(request):
    departments = Department.objects.all().order_by('id')
    create_form = DepartmentForm()  # for Add modal  
    department_with_forms = [(department, DepartmentForm(instance=department)) for department in departments]
    return render(request, 'department_list.html', {'departments': departments, 'create_form': create_form, 'department_with_forms': department_with_forms})

def department_create(request):
    if request.method == 'POST':
        form = DepartmentForm(request.POST)
        if form.is_valid():
            form.save()
    return redirect('department_list')
def department_edit(request, department_id):
    department = get_object_or_404(Department, pk=department_id)
    if request.method == 'POST':
        form = DepartmentForm(request.POST, instance=department)
        if form.is_valid():
            form.save()
    return redirect('department_list')

def department_delete(request, department_id):
    department = get_object_or_404(Department, pk=department_id)
    department.delete()
    return redirect('department_list')

def main_menu(request):
    return render(request, 'main_menu.html')
 
 